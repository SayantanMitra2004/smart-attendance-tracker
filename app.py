from flask import Flask, render_template, Response, request, redirect, url_for, jsonify, flash, send_file, session
from flask_cors import CORS
from dotenv import load_dotenv
import os
import time
load_dotenv()
import pandas as pd
import cv2
import pickle
import face_recognition
import numpy as np
import EncodeGenerator
import json
import io
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

from threading import Thread
import sqlite3

app = Flask(__name__, static_folder = "./static", static_url_path="/")
CORS(app)
app.secret_key = 'supersecretkey'  

UPLOAD_FOLDER = 'uploads'
USER_DATA_FILE = 'user_data.json'


DATABASE = 'users.db'


app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

#user database function
def init_db():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            user_id TEXT PRIMARY KEY,
            user_name TEXT,
            department TEXT,
            semester TEXT
        )
    ''')
    conn.commit()
    conn.close()

#attendance database function
def init_attendance_db():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id TEXT,
            timestamp TEXT,         
            is_present BOOLEAN,
            FOREIGN KEY(user_id) REFERENCES users(user_id)
        )
    ''')
    
    cursor.execute("SELECT user_id FROM users")
    user_ids = cursor.fetchall()
    

    # 4. Inserting user_id into table if not already present
    for (user_id,) in user_ids:
        cursor.execute(f'SELECT 1 FROM attendance WHERE user_id = ?', (user_id,))
        if not cursor.fetchone():
            cursor.execute(f'INSERT INTO attendance (user_id) VALUES (?)', (user_id,))
        # Set today's date column to False (0)
        cursor.execute(f'UPDATE attendance SET is_present = 0 WHERE user_id = ?', (user_id,))


    conn.commit()
    conn.close()

 #calling criteria
def should_run_today(date_file_path='last_run.txt'):
    curr_date = datetime.now().strftime('%Y-%m-%d')
    if os.path.exists(date_file_path):
        with open(date_file_path, 'r') as file:
            last_run = file.read().strip()
            if last_run == curr_date:
                return False  # Already ran today
    with open(date_file_path, 'w') as file:
        file.write(curr_date)
    return True

# calling once every day
if should_run_today():
    init_db()
    init_attendance_db()

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# WEB CAM VIDEO CAPTURE
camera = cv2.VideoCapture(0)

# LOADING ENCODED FILE AND SAVE IT AS A PICKLE FILE
print("Loading Encoded File...")
file = open('Encodefile.p', 'rb')
encodeListKnownWithIds = pickle.load(file)
file.close()

encodeListKnown, studentIds = encodeListKnownWithIds
print("Encoded file loaded...")
print(studentIds)


# Global variable to track face recognition
face_recognized = False
recognized_user_id = None  # Add this as a global variable

# Dictionary to store room deadlines
room_deadlines = {}


def generate_frames():
    global face_recognized, studentIds, recognized_user_id
    while True:
        success, frame = camera.read()
        if not success:
            break
        else:
            imgS = cv2.resize(frame, (0, 0), None, 0.25, 0.25)
            imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

            faceCurFrame = face_recognition.face_locations(imgS)
            encodeCurFrame = face_recognition.face_encodings(imgS, faceCurFrame)

            if faceCurFrame:
                for encodeFace, faceloc in zip(encodeCurFrame, faceCurFrame):
                    matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
                    faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)

                    matchIndex = np.argmin(faceDis)
                    if matches[matchIndex]:
                        recognized_user_id = studentIds[matchIndex]
                        face_recognized = True
                        break  # Exit loop once face is recognized

            ret, buffer = cv2.imencode('.jpg', frame)
            frame = buffer.tobytes()

            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')


@app.route('/', methods=['GET', 'POST'])
def home():
    return render_template('home.html')


@app.route('/index', methods=['GET', 'POST'])
def index():
    global face_recognized, recognized_user_id
    if face_recognized and recognized_user_id:
        face_recognized = False  # Reset for next session

    return render_template('index.html')


@app.route('/signup')
def signup():
    return render_template('signup.html')

@app.route('/login')
def login():
    return render_template('login.html')

@app.route('/video')
def video():
    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/admin')
def admin():
    return render_template('admin.html')

@app.route('/delete_entry_page')
def delete_entry_page():
   return render_template('delete_entry_page.html')

@app.route('/delete_attendance_page')
def delete_attendance_page():
   return render_template('delete_attendance_page.html')

@app.route('/admin_login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        admin_id = request.form['admin_id']
        password = request.form['password']

        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM admin WHERE admin_id=? AND password=?", 
                       (admin_id, password))
        admin = cursor.fetchone()
        conn.close()

        if admin:
            return redirect(url_for('admin'))
        else:
            error = "Invalid credentials. Please try again."
            return render_template('admin_login.html', error=error)

    return render_template('admin_login.html')

@app.route('/modify_admin', methods=['GET', 'POST'])
def modify_admin():
    if request.method == 'POST':
        admin_id = request.form['admin_id']
        password = request.form['password']

        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()

        # Check if an admin already exists
        cursor.execute("SELECT COUNT(*) FROM admin")
        count = cursor.fetchone()[0]

        if count == 0:
            # Insert if table is empty
            cursor.execute("INSERT INTO admin (admin_id, password) VALUES (?, ?)", (admin_id, password))
        else:
            # Update the existing admin
            cursor.execute("UPDATE admin SET admin_id = ?, password = ?", (admin_id, password))

        conn.commit()
        conn.close()
        flash("Admin Modified Successfully", "success")
        return redirect('/modify_admin')  # or wherever you want to redirect after update

    return render_template('modify_admin.html')  # Create a simple form to update admin


@app.route('/teacher_login', methods=['GET', 'POST'])
def teacher_login():
    if request.method == 'POST':
        teacher_id = request.form['teacher_id']
        teacher_name = request.form['teacher_name']
        password = request.form['password']

        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM teachers WHERE teacher_id=? AND teacher_name=? AND password=?", 
                       (teacher_id, teacher_name, password))
        teacher = cursor.fetchone()
        conn.close()

        if teacher:
            return redirect(url_for('view_attendance'))
        else:
            error = "Invalid credentials. Please try again."
            return render_template('teacher_login.html', error=error)

    return render_template('teacher_login.html')


@app.route('/register_teacher', methods=['GET', 'POST'])
def register_teacher():
    if request.method == 'POST':
        teacher_id = request.form.get('teacher_id')
        teacher_name = request.form.get('teacher_name')
        password = request.form.get('password')
        department = request.form.get('department')
        designation = request.form.get('designation')
        year_of_joining = request.form.get('joined_year')

        if not all([teacher_id, teacher_name, password, department, designation, year_of_joining]):
            flash("All fields are required!", "error")
            return redirect('/register_teacher')

        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()

        # Create table if not exists
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS teachers (
                teacher_id TEXT PRIMARY KEY,
                teacher_name TEXT,
                password TEXT,
                department TEXT,
                designation TEXT,
                year_of_joining TEXT
            )
        ''')

        # Insert teacher if not exists
        cursor.execute("SELECT * FROM teachers WHERE teacher_id = ?", (teacher_id,))
        existing_teacher = cursor.fetchone()

        if existing_teacher:
            flash("Teacher already exists!", "error")
            conn.close()
            return redirect('/register_teacher')

        cursor.execute('''
            INSERT INTO teachers (teacher_id, teacher_name, password, department, designation, year_of_joining)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (teacher_id, teacher_name, password, department, designation, year_of_joining))

        conn.commit()
        conn.close()

        flash("Teacher registered successfully!", "success")
        return redirect('/register_teacher')

    return render_template('register_teacher.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    form_values = request.form.to_dict()
    user_id = form_values["user_id"]
    user_name = form_values["user_name"]
    department = form_values["department"]
    semester = form_values["semester"]

    if 'image' not in request.files:
        return 'No file part'

    file = request.files['image']

    if file.filename == '':
        return 'No selected file'

    if file:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM users WHERE user_id = ?", (user_id,))
        existing_user = cursor.fetchone()

        if not existing_user:
            cursor.execute(
                "INSERT INTO users (user_id, user_name, department, semester) VALUES (?, ?, ?, ?)",
                (user_id, user_name, department, semester)
            )
            conn.commit()
            conn.close()

        
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], user_id + ".png")
            file.save(file_path)
            
            # Regenerate encodings
            EncodeGenerator.encodeGenerator()

            return redirect('/')
        else:
            conn.close()
            flash("User already exists!", "error")
            return redirect('/signup')  

    return 'File upload failed'



@app.route('/check_face')
def check_face():
    global face_recognized, recognized_user_id

    if face_recognized and recognized_user_id:
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()

        # Check if user exists
        cursor.execute("SELECT user_name, department, semester FROM users WHERE user_id = ?", (recognized_user_id,))
        result = cursor.fetchone()
        print(recognized_user_id)

        if not result:
            conn.close()
            return jsonify({
                'recognized': True,
                'user_id': recognized_user_id,
                'status': 'not_found',
                'message': 'User not found in database.'
            })

        user_name, department, semester = result

        # Check if user has already been marked present today
        today_date = datetime.now().strftime('%Y-%m-%d')
        cursor.execute("""
            SELECT 1 FROM attendance 
            WHERE user_id = ? AND DATE(timestamp) = ? AND is_present = 1
        """, (recognized_user_id, today_date))

        if cursor.fetchone():
            conn.close()
            return jsonify({
                'recognized': True,
                'user_id': recognized_user_id,
                'status': 'duplicate',
                'message': 'Attendance already marked for today.'
            })

        # Mark as present
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute("""
            UPDATE attendance
            SET is_present = 1, timestamp = ?
            WHERE user_id = ?
        """, (timestamp, recognized_user_id))

        conn.commit()
        conn.close()

        return jsonify({
            'recognized': True,
            'user_id': recognized_user_id,
            'status': 'success',
            'message': 'Attendance marked successfully.'
        })

    return jsonify({
        'recognized': False,
        'user_id': None,
        'status': 'not_recognized',
        'message': 'Face not recognized.'
    })


@app.route('/stop_camera')
def stop_camera():
    global camera
    camera.release()
    return jsonify({'status': 'camera stopped'})


@app.route('/start_camera')
def start_camera():
    global camera
    if camera is None or not camera.isOpened():
        camera = cv2.VideoCapture(0)
        return jsonify({'status': 'camera started'})
    else:
        return jsonify({'status': 'camera already running'})


@app.route('/attendance', methods=['GET', 'POST']) 
def view_attendance():
    if request.method == 'POST':
        subject_name = request.form['subject_name'].strip().replace(" ", "_").lower()
        date_column = request.form['date_column'].strip()
        semester = request.form['semester'].strip()
        department = request.form['department'].strip()
        table_name = f'attendance_{subject_name}_{semester}_{department}'

        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()

        # 1. Create subject-wise table if not exists
        cursor.execute(f'''
            CREATE TABLE IF NOT EXISTS "{table_name}" (
                user_id TEXT UNIQUE,
                FOREIGN KEY(user_id) REFERENCES users(user_id)
            )
        ''')

        
        # 2. Add new date column if not exists
        cursor.execute(f'PRAGMA table_info("{table_name}")')
        existing_columns = [col[1] for col in cursor.fetchall()]
        new_column_added = False
        if date_column not in existing_columns:
            cursor.execute(f'''
                ALTER TABLE "{table_name}"
                ADD COLUMN "{date_column}" BOOLEAN DEFAULT 0
            ''')
            new_column_added = True


        # 3. Fetch users for the given semester and department
        cursor.execute("SELECT user_id FROM users WHERE semester = ? and department = ?", (semester, department))
        user_ids = cursor.fetchall()

        
        # 4. Insert new user_ids and set initial attendance only if column was newly added
        for (user_id,) in user_ids:
            cursor.execute(f'INSERT OR IGNORE INTO "{table_name}" (user_id) VALUES (?)', (user_id,))
            if new_column_added:
                cursor.execute(f'''
                    UPDATE "{table_name}" SET "{date_column}" = 0 WHERE user_id = ?
                ''', (user_id,))


        
        conn.commit()

        # 5. Fetch attendance data for display
        cursor.execute(f'''
            SELECT users.user_id, users.user_name, users.department, users.semester, 
                   attendance.is_present, attendance.timestamp,
                   "{table_name}"."{date_column}"
            FROM users
            LEFT JOIN attendance ON users.user_id = attendance.user_id
            LEFT JOIN "{table_name}" ON users.user_id = "{table_name}".user_id 
            WHERE users.semester = ? and users.department = ?
        ''', (semester,department))
        rows = cursor.fetchall()
        print(rows)
        conn.close()

        return render_template('attendance.html', rows=rows, subject_name=subject_name, date=date_column, department=department, semester=semester)

    return render_template('attendance.html')


@app.route('/delete_subject_attendance/<subject_name>', methods=['POST'])
def delete_subject_attendance(subject_name):
    try:
        # Sanitize subject_name
        subject_name = subject_name.strip()

        if not subject_name:
            return jsonify({'error': 'Subject name cannot be empty.'}), 400

        data = request.get_json()
        semester = data.get('semester', '').strip()
        department = data.get('department', '').strip()

        if not semester:
            return jsonify({'error': 'Semester is required.'}), 400
        if not department:
            return jsonify({'error': 'Department is required.'}), 400

        # You can customize your table naming or filtering logic based on semester and department here
        # Example: You might want to include semester and department in your table name or your deletion criteria

        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()

        # For example, table naming includes subject name only (modify if needed)
        table_name = f'attendance_{subject_name}_{semester}_{department}'

        # Check if the subject table exists
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
        if not cursor.fetchone():
            conn.close()
            return jsonify({'error': f"Subject '{subject_name}' does not exist."}), 404

        # Example: If you want to delete attendance only for given semester and department inside the table,
        # you would do a DELETE query here instead of dropping the whole table.
        # But since your code drops the whole table, I keep it as is.

        cursor.execute(f"DROP TABLE IF EXISTS '{table_name}'")
        conn.commit()
        conn.close()

        return jsonify({'message': f"Subject '{subject_name}' attendance data deleted successfully for semester '{semester}' and department '{department}'."})

    except Exception as e:
        print("Error:", e)
        return jsonify({'error': 'An error occurred while deleting the subject attendance.'}), 500



@app.route('/delete_entry/<student_id>', methods=['POST'])
def delete_entry(student_id):
    import EncodeGenerator
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    try:
        # Remove encoding
        success = EncodeGenerator.removeEncoding(student_id)

        if not success:
            return jsonify({"error": f"Student ID {student_id} not found in encodings"}), 404

        # Delete from users table
        cursor.execute("DELETE FROM users WHERE user_id = ?", (student_id,))

        #  Delete from attendance table
        cursor.execute("DELETE FROM attendance WHERE user_id = ?", (student_id,))

        # Delete from all subject-wise attendance tables
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'attendance_%'")
        subject_tables = cursor.fetchall()

        for table in subject_tables:
            cursor.execute(f"DELETE FROM {table[0]} WHERE user_id = ?", (student_id,))

        # Delete image file from uploads folder
        image_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{student_id}.png")
        if os.path.exists(image_path):
            os.remove(image_path)

        # Re-generate encodings after deletion
        EncodeGenerator.encodeGenerator()

        conn.commit()
        return jsonify({"message": f"Successfully deleted {student_id} from database, encodings, and image files"}), 200

    except Exception as e:
        conn.rollback()
        return jsonify({"error": f"Error deleting student: {str(e)}"}), 500

    finally:
        conn.close()


@app.route('/submit_attendance', methods=['POST'])
def submit_attendance():
    selected_ids = request.form.getlist('selected_ids')
    subject_name = request.form.get('subject_name').strip()
    date_column = request.form.get('date_column').strip()
    department = request.form.get('department')
    semester = request.form.get('semester')
    print(semester)
    print(department)
    if not subject_name or not date_column:
        return "Missing subject name or date column", 400

    subject_table_name = f'attendance_{subject_name}_{semester}_{department}'

    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()

    # print("Selected IDs:", selected_ids)
    print("Subject Table:", subject_table_name)
    # print("Date Column:", date_column)

    try:
        # Mark everyone absent
        cursor.execute(f'''
            UPDATE {subject_table_name}
            SET "{date_column}" = 0
        ''')

        # Then mark only selected users as present
        for user_id in selected_ids:
            cursor.execute(f'''
                UPDATE {subject_table_name}
                SET "{date_column}" = 1
                WHERE user_id = ?
            ''', (user_id,))
            print(f"Marked present: {user_id}")

    except Exception as e:
        print("Error during attendance update:", e)

    conn.commit()
    conn.close()

    return redirect('/attendance')

from datetime import datetime

@app.route('/download_subject_attendance', methods=['GET', 'POST'])
def download_subject_attendance():
    if request.method == 'POST':
        subject_name = request.form.get('subject_name')
        semester = request.form.get('semester')
        department = request.form.get('department')
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')

        

        if not subject_name or not start_date or not end_date or not department or not semester:
            return "Subject, Department, Semester, Start and End dates are required", 400


        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()

        subject_table_name = f"attendance_{subject_name}_{semester}_{department}"

        try:
            # Get all column names from the subject table
            cursor.execute(f"PRAGMA table_info('{subject_table_name}')")
            columns_info = cursor.fetchall()
            column_names = [col[1] for col in columns_info if col[1] not in ('id', 'user_id')]

            # Filter columns based on date range
            filtered_columns = [col for col in column_names if start_date <= col <= end_date]

            if not filtered_columns:
                flash("⚠️ No attendance data available for the selected date range.", "warning")
                return redirect(url_for('download_subject_attendance'))


            # Create SQL query string for filtered date columns
            column_selection = ', '.join([f's."{col}"' for col in filtered_columns])

            # SQL query with proper alias for student_id
            sql = f'''
                SELECT u.user_id AS student_id, u.user_name, u.semester, {column_selection}
                FROM users u
                INNER JOIN "{subject_table_name}" s ON u.user_id = s.user_id
                WHERE u.semester IN (SELECT DISTINCT semester FROM users)
            '''

            df = pd.read_sql_query(sql, conn)

            # Save original numerical values for color formatting later
            original_values = df[filtered_columns].copy()

            # Replace 1 -> ✅ and 0 -> ❌ in only the attendance columns
            for col in filtered_columns:
                df[col] = df[col].apply(lambda x: '✅' if x == 1 else ('❌' if x == 0 else x))


            # Output Excel file
            output_filename = f'{subject_name}_attendance_{start_date}_to_{end_date}.xlsx'
            output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
            df.to_excel(output_path, index=False)

            # Open workbook to apply fill color
            wb = load_workbook(output_path)
            ws = wb.active

            # Color fill definitions
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            # Apply color fills based on original numeric values
            for row_idx, row in original_values.iterrows():
                for col_idx, col in enumerate(filtered_columns, start=4):  
                    val = row[col]
                    cell = ws.cell(row=row_idx + 2, column=col_idx) 
                    if val == 1:
                        cell.fill = green_fill
                    elif val == 0:
                        cell.fill = red_fill

            wb.save(output_path)
            
            conn.close()
            response =  send_file(output_path, as_attachment=True)

            # Remove the file after download is complete )
            def delete_file_later():
                # Delay in seconds before file deletion 
                time.sleep(5)
                os.remove(output_path)
            
            # Run the deletion in the background
            
            Thread(target=delete_file_later).start()

            return response

        except sqlite3.Error as e:
            conn.close()
            return f"Database error: {str(e)}", 500

    return render_template('download_subject_attendance.html')

if __name__ == "__main__":
    app.run(debug=True)